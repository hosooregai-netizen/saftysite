#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


CAUSATIVE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("12_이동식크레인", ("이동식 크레인", "슬링벨트")),
    ("10_이동식비계", ("이동식 비계",)),
    ("11_거푸집동바리", ("동바리", "거푸집", "갱폼")),
    ("6_고소작업대", ("차량계 고소작업대", "고소작업대")),
    ("7_사다리", ("승강용 사다리", "이동식 사다리", "사다리")),
    ("8_달비계", ("달비계",)),
    ("1_단부_개구부", ("개구부", "단부", "엘리베이터실", "계단 단부")),
    ("3_지붕", ("지붕", "채광창")),
    ("2_철골", ("철골", "h-빔", "h빔", "철골자재")),
    ("9_트럭", ("덤프트럭", "트럭", "적재함", "지게차", "집게차", "사다리차", "펌프카")),
    ("5_굴착기", ("미니굴착기", "굴착기", "천공기", "항타기", "노면파쇄기")),
    ("4_비계_작업발판", ("시스템 비계", "작업발판", "말비계", "비계")),
    ("13_화재_폭발", ("화상", "화재", "폭발")),
]


ACCIDENT_TYPE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("추락", ("추락",)),
    ("떨어짐", ("떨어짐",)),
    ("끼임", ("끼임",)),
    ("부딪힘", ("부딪힘",)),
    ("맞음", ("맞음",)),
    ("깔림", ("깔림",)),
    ("매몰", ("매몰", "매몰2")),
    ("감전", ("감전",)),
    ("화상", ("화상",)),
    ("찔림", ("찔림",)),
]


def normalize_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\xa0", " ")).strip()


def normalize_slug(value: str | None) -> str:
    normalized = normalize_text(value).lower()
    normalized = normalized.replace(",", " ")
    return re.sub(r"[^0-9a-z가-힣]+", "-", normalized).strip("-")


def split_body(raw: str | None) -> tuple[str, str]:
    text = (raw or "").replace("\r\n", "\n")
    matched = re.search(r"재해내용\s*:\s*(.*?)\s*예방대책\s*:\s*(.*)", text, re.S)
    if matched:
      return normalize_text(matched.group(1)), normalize_text(matched.group(2))

    chunks = [normalize_text(chunk) for chunk in text.split("\n\n") if normalize_text(chunk)]
    if not chunks:
        return "", ""
    if len(chunks) == 1:
        return chunks[0], ""
    return chunks[0], " ".join(chunks[1:])


def parse_title(raw: str | None) -> tuple[str, str, str]:
    clean = " ".join(str(raw or "").split())
    matched = re.match(r"(.+?)\s*\((.+)\)$", clean)
    if matched:
        return clean, matched.group(1).strip(), matched.group(2).strip()
    return clean, clean, ""


def infer_causative_key(*parts: str) -> str:
    haystack = normalize_text(" ".join(parts)).lower()
    for key, keywords in CAUSATIVE_RULES:
        if any(keyword.lower() in haystack for keyword in keywords):
            return key
    return "14_기타_위험요인"


def infer_accident_type(*parts: str) -> str:
    haystack = normalize_text(" ".join(parts)).lower()
    for accident_type, keywords in ACCIDENT_TYPE_RULES:
        if any(keyword.lower() in haystack for keyword in keywords):
            return accident_type
    return "기타"


def build_search_keywords(*parts: str) -> list[str]:
    tokens: list[str] = []
    for part in parts:
        for token in re.split(r"[^0-9A-Za-z가-힣]+", normalize_text(part)):
            if len(token) < 2:
                continue
            if token not in tokens:
                tokens.append(token)
    return tokens[:24]


def find_workbook(root: Path) -> Path:
    workbooks = sorted(root.rglob("*.xlsx"))
    if not workbooks:
        raise FileNotFoundError("zip 내부에서 xlsx 파일을 찾지 못했습니다.")
    return workbooks[0]


def find_image_path(root: Path, folder: str, index: int) -> str:
    target_dir = root / folder
    if not target_dir.exists():
        return ""

    prefix = f"{index}. "
    matches = sorted(
        path for path in target_dir.iterdir() if path.is_file() and path.name.startswith(prefix)
    )
    if not matches:
        return ""

    return str(matches[0].relative_to(root))


def parse_catalog(zip_path: Path) -> list[dict[str, object]]:
    with tempfile.TemporaryDirectory(prefix="disaster_case_catalog_") as temp_dir:
        extract_root = Path(temp_dir)
        with zipfile.ZipFile(zip_path) as archive:
            archive.extractall(extract_root)

        workbook_path = find_workbook(extract_root)
        workbook = load_workbook(workbook_path, data_only=True)

        entries: list[dict[str, object]] = []
        sequence = 1

        for sheet in workbook.worksheets:
            sheet_category = sheet.title.split("-")[-1]

            for row in sheet.iter_rows(values_only=True):
                for base_index in (1, 8):
                    index = row[base_index] if len(row) > base_index else None
                    title = row[base_index + 1] if len(row) > base_index + 1 else None
                    body = row[base_index + 2] if len(row) > base_index + 2 else None
                    if not index or not title or not body:
                        continue

                    clean_title, accident_category, detail_label = parse_title(str(title))
                    description, prevention_measure = split_body(str(body))
                    image_path = find_image_path(extract_root, sheet_category, int(index))
                    recommended_accident_type = infer_accident_type(
                        accident_category, detail_label, description
                    )
                    recommended_causative_key = infer_causative_key(
                        detail_label, description, prevention_measure
                    )

                    entries.append(
                        {
                            "id": f"disaster-case-{sequence:03d}",
                            "sheetName": sheet.title,
                            "sheetCategory": sheet_category,
                            "index": int(index),
                            "title": clean_title,
                            "accidentCategory": accident_category,
                            "detailLabel": detail_label,
                            "description": description,
                            "preventionMeasure": prevention_measure,
                            "imagePath": image_path,
                            "recommendedAccidentType": recommended_accident_type,
                            "recommendedCausativeKey": recommended_causative_key,
                            "searchKeywords": build_search_keywords(
                                clean_title, description, prevention_measure
                            ),
                        }
                    )
                    sequence += 1

        return entries


def main() -> None:
    parser = argparse.ArgumentParser(description="재해사례 zip/xlsx를 JSON 카탈로그로 변환합니다.")
    parser.add_argument("zip_path", type=Path, help="원본 재해사례 zip 파일 경로")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("data/disaster-case-catalog.json"),
        help="생성할 JSON 경로",
    )
    args = parser.parse_args()

    catalog = parse_catalog(args.zip_path)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"generated {len(catalog)} entries -> {args.out}")


if __name__ == "__main__":
    main()
