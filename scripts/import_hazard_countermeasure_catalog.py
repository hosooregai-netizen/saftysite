#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_SAFETY_API_UPSTREAM_BASE_URL = "http://52.64.85.49:8011/api/v1"
API_BASE_ENV_KEYS = (
    "SAFETY_API_UPSTREAM_BASE_URL",
    "SAFETY_API_BASE_URL",
    "NEXT_PUBLIC_SAFETY_API_UPSTREAM_BASE_URL",
    "NEXT_PUBLIC_SAFETY_API_BASE_URL",
)
CONTENT_LIST_LIMIT = 1000
REQUIRED_ROW_COUNT_PER_SHEET = 161
EXPECTED_TOTAL_ROWS = REQUIRED_ROW_COUNT_PER_SHEET * 2
EXPECTED_SHEET1_NONEMPTY_CATEGORY_COUNT = 25


@dataclass(frozen=True)
class SheetConfig:
    title: str
    code_prefix: str
    sheet_tag: str
    category_index: int | None
    title_index: int
    expected_risk_index: int
    countermeasure_index: int
    legal_reference_index: int
    note_index: int
    expected_header: tuple[str, ...]


@dataclass(frozen=True)
class ImportRow:
    source_sheet: str
    source_row_number: int
    sequence_in_sheet: int
    code: str
    sort_order: int
    title: str
    category: str
    expected_risk: str
    countermeasure: str
    legal_reference: str
    note: str
    tags: tuple[str, ...]

    def to_payload(self) -> dict[str, Any]:
        return {
            "content_type": "hazard_countermeasure_catalog",
            "title": self.title,
            "code": self.code,
            "body": {
                "category": self.category,
                "expectedRisk": self.expected_risk,
                "countermeasure": self.countermeasure,
                "legalReference": self.legal_reference,
                "note": self.note,
            },
            "tags": list(self.tags),
            "sort_order": self.sort_order,
            "effective_from": None,
            "effective_to": None,
            "is_active": True,
        }


SHEET_CONFIGS = (
    SheetConfig(
        title="Table 1",
        code_prefix="hazard-cm-table1",
        sheet_tag="sheet-table1",
        category_index=None,
        title_index=1,
        expected_risk_index=2,
        countermeasure_index=3,
        legal_reference_index=4,
        note_index=5,
        expected_header=(
            "no",
            "제목",
            "예상위험",
            "관리대책",
            "법령",
            "비고(#3,단부,개부부,발판)",
        ),
    ),
    SheetConfig(
        title="Sheet1",
        code_prefix="hazard-cm-sheet1",
        sheet_tag="sheet-sheet1",
        category_index=1,
        title_index=2,
        expected_risk_index=3,
        countermeasure_index=4,
        legal_reference_index=5,
        note_index=6,
        expected_header=(
            "no",
            "구분",
            "제목",
            "예상위험",
            "관리대책",
            "법령",
            "비고(특징)",
        ),
    ),
)


def normalize_base_url(value: str) -> str:
    return value.rstrip("/")


def is_absolute_http_url(value: str) -> bool:
    return re.match(r"^https?://", value, flags=re.I) is not None


def resolve_api_base_url(explicit: str | None) -> str:
    if explicit and is_absolute_http_url(explicit.strip()):
        return normalize_base_url(explicit.strip())

    for env_key in API_BASE_ENV_KEYS:
        configured = os.getenv(env_key, "").strip()
        if configured and is_absolute_http_url(configured):
            return normalize_base_url(configured)

    return DEFAULT_SAFETY_API_UPSTREAM_BASE_URL


def normalize_line_endings(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_line_endings(value).strip()
    return normalize_line_endings(str(value)).strip()


def normalize_match_text(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_line_endings(value)).strip()


def normalize_header_cell(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_cell_text(value))


def request_json(
    method: str,
    url: str,
    *,
    headers: dict[str, str] | None = None,
    body: bytes | None = None,
    timeout: int = 30,
) -> Any:
    request = Request(url, data=body, method=method.upper())
    for key, value in (headers or {}).items():
        request.add_header(key, value)

    try:
        with urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
            if not raw.strip():
                return None
            return json.loads(raw)
    except HTTPError as error:
        raw = error.read().decode("utf-8", errors="replace")
        detail = raw
        if raw:
            try:
                payload = json.loads(raw)
                if isinstance(payload, dict) and isinstance(payload.get("detail"), str):
                    detail = payload["detail"]
                else:
                    detail = json.dumps(payload, ensure_ascii=False)
            except json.JSONDecodeError:
                detail = raw
        raise RuntimeError(f"{method.upper()} {url} failed ({error.code}): {detail}") from error
    except URLError as error:
        raise RuntimeError(f"{method.upper()} {url} failed: {error.reason}") from error


def resolve_access_token(api_base_url: str) -> str:
    access_token = os.getenv("SAFETY_ACCESS_TOKEN", "").strip()
    if access_token:
        return access_token

    email = os.getenv("SAFETY_ADMIN_EMAIL", "").strip()
    password = os.getenv("SAFETY_ADMIN_PASSWORD")
    if not email or password is None:
        raise RuntimeError(
            "SAFETY_ACCESS_TOKEN 또는 SAFETY_ADMIN_EMAIL / SAFETY_ADMIN_PASSWORD 환경변수가 필요합니다."
        )

    body = urlencode({"username": email, "password": password}).encode("utf-8")
    response = request_json(
        "POST",
        f"{api_base_url}/auth/token",
        headers={"Content-Type": "application/x-www-form-urlencoded; charset=utf-8"},
        body=body,
    )
    token = response.get("access_token") if isinstance(response, dict) else None
    if not isinstance(token, str) or not token.strip():
        raise RuntimeError("인증 응답에서 access_token을 찾지 못했습니다.")
    return token.strip()


def fetch_all_content_items(
    api_base_url: str,
    access_token: str,
    *,
    active_only: bool,
    include_body: bool,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    offset = 0
    headers = {"Authorization": f"Bearer {access_token}"}

    while True:
        query = urlencode(
            {
                "active_only": str(active_only).lower(),
                "include_body": str(include_body).lower(),
                "limit": CONTENT_LIST_LIMIT,
                "offset": offset,
            }
        )
        page = request_json(
            "GET",
            f"{api_base_url}/content-items?{query}",
            headers=headers,
        )
        if not isinstance(page, list):
            raise RuntimeError("content-items 응답 형식이 예상과 다릅니다.")

        items.extend(page)
        if len(page) < CONTENT_LIST_LIMIT:
            return items
        offset += len(page)


def post_content_item(api_base_url: str, access_token: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = request_json(
        "POST",
        f"{api_base_url}/content-items",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        body=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
    )
    if not isinstance(response, dict):
        raise RuntimeError("content-items 생성 응답 형식이 예상과 다릅니다.")
    return response


def iter_nonempty_rows(worksheet: Any) -> list[tuple[int, tuple[Any, ...]]]:
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if any(normalize_cell_text(cell) for cell in row):
            rows.append((row_number, row))
    return rows


def validate_sheet_header(config: SheetConfig, header_row: tuple[Any, ...]) -> None:
    normalized_actual = tuple(normalize_header_cell(cell) for cell in header_row)
    normalized_expected = tuple(normalize_header_cell(cell) for cell in config.expected_header)
    if normalized_actual[: len(normalized_expected)] != normalized_expected:
        raise RuntimeError(
            f"{config.title} 헤더가 예상과 다릅니다.\n"
            f"expected={normalized_expected}\nactual={normalized_actual}"
        )


def collect_sheet_rows(worksheet: Any, config: SheetConfig) -> list[tuple[int, tuple[Any, ...]]]:
    nonempty_rows = iter_nonempty_rows(worksheet)
    if not nonempty_rows:
        raise RuntimeError(f"{config.title} 시트가 비어 있습니다.")

    _, header_row = nonempty_rows[0]
    validate_sheet_header(config, header_row)

    data_rows = nonempty_rows[1:]
    if len(data_rows) != REQUIRED_ROW_COUNT_PER_SHEET:
        raise RuntimeError(
            f"{config.title} 데이터 행 수가 예상과 다릅니다. "
            f"expected={REQUIRED_ROW_COUNT_PER_SHEET}, actual={len(data_rows)}"
        )
    return data_rows


def build_sheet1_legal_lookup(sheet_rows: list[tuple[int, tuple[Any, ...]]], config: SheetConfig) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for _, row in sheet_rows:
        title = normalize_cell_text(row[config.title_index])
        legal_reference = normalize_cell_text(row[config.legal_reference_index])
        if title and legal_reference:
            lookup.setdefault(normalize_match_text(title), legal_reference)
    return lookup


def build_import_rows(workbook_path: Path) -> list[ImportRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}

    for config in SHEET_CONFIGS:
        if config.title not in worksheets:
            raise RuntimeError(f"{config.title} 시트를 찾지 못했습니다.")

    sheet_rows_by_name = {
        config.title: collect_sheet_rows(worksheets[config.title], config)
        for config in SHEET_CONFIGS
    }
    sheet1_lookup = build_sheet1_legal_lookup(
        sheet_rows_by_name["Sheet1"],
        next(config for config in SHEET_CONFIGS if config.title == "Sheet1"),
    )

    combined_rows: list[ImportRow] = []
    sort_order = 0

    for config in SHEET_CONFIGS:
        for sequence_in_sheet, (source_row_number, row) in enumerate(
            sheet_rows_by_name[config.title], start=1
        ):
            title = normalize_cell_text(row[config.title_index])
            expected_risk = normalize_cell_text(row[config.expected_risk_index])
            countermeasure = normalize_cell_text(row[config.countermeasure_index])
            legal_reference = normalize_cell_text(row[config.legal_reference_index])
            note = normalize_cell_text(row[config.note_index])
            category = (
                normalize_cell_text(row[config.category_index])
                if config.category_index is not None
                else ""
            )

            if config.title == "Table 1" and not legal_reference:
                legal_reference = sheet1_lookup.get(normalize_match_text(title), "")

            combined_rows.append(
                ImportRow(
                    source_sheet=config.title,
                    source_row_number=source_row_number,
                    sequence_in_sheet=sequence_in_sheet,
                    code=f"{config.code_prefix}-{sequence_in_sheet:03d}",
                    sort_order=sort_order,
                    title=title,
                    category=category,
                    expected_risk=expected_risk,
                    countermeasure=countermeasure,
                    legal_reference=legal_reference,
                    note=note,
                    tags=("excel-import", "hazard-countermeasure", config.sheet_tag),
                )
            )
            sort_order += 1

    if len(combined_rows) != EXPECTED_TOTAL_ROWS:
        raise RuntimeError(
            f"변환된 총 행 수가 예상과 다릅니다. expected={EXPECTED_TOTAL_ROWS}, actual={len(combined_rows)}"
        )

    return combined_rows


def summarize_import_rows(rows: list[ImportRow]) -> dict[str, Any]:
    table1_rows = [row for row in rows if row.source_sheet == "Table 1"]
    sheet1_rows = [row for row in rows if row.source_sheet == "Sheet1"]
    required_blank_counts = {
        "title": sum(1 for row in rows if not row.title),
        "expectedRisk": sum(1 for row in rows if not row.expected_risk),
        "countermeasure": sum(1 for row in rows if not row.countermeasure),
        "legalReference": sum(1 for row in rows if not row.legal_reference),
        "note": sum(1 for row in rows if not row.note),
    }
    sheet1_nonempty_category_count = sum(1 for row in sheet1_rows if row.category)
    fallback_row = next(
        row
        for row in table1_rows
        if row.code == "hazard-cm-table1-012"
    )
    return {
        "table1_rows": len(table1_rows),
        "sheet1_rows": len(sheet1_rows),
        "total_rows": len(rows),
        "sheet1_nonempty_category_count": sheet1_nonempty_category_count,
        "required_blank_counts": required_blank_counts,
        "fallback_row_legal_reference": fallback_row.legal_reference,
    }


def validate_import_rows(rows: list[ImportRow]) -> dict[str, Any]:
    summary = summarize_import_rows(rows)
    if summary["table1_rows"] != REQUIRED_ROW_COUNT_PER_SHEET:
        raise RuntimeError("Table 1 행 수가 161이 아닙니다.")
    if summary["sheet1_rows"] != REQUIRED_ROW_COUNT_PER_SHEET:
        raise RuntimeError("Sheet1 행 수가 161이 아닙니다.")
    if summary["total_rows"] != EXPECTED_TOTAL_ROWS:
        raise RuntimeError("총 행 수가 322가 아닙니다.")
    if summary["sheet1_nonempty_category_count"] != EXPECTED_SHEET1_NONEMPTY_CATEGORY_COUNT:
        raise RuntimeError(
            "Sheet1 구분 값 개수가 예상과 다릅니다. "
            f"expected={EXPECTED_SHEET1_NONEMPTY_CATEGORY_COUNT}, "
            f"actual={summary['sheet1_nonempty_category_count']}"
        )
    blank_counts = summary["required_blank_counts"]
    nonzero_blanks = {key: value for key, value in blank_counts.items() if value}
    if nonzero_blanks:
        raise RuntimeError(f"필수 필드 공란이 남아 있습니다: {nonzero_blanks}")
    if summary["fallback_row_legal_reference"] != "산업안전보건기준에 관한 규칙 제42조(추락의 방지)":
        raise RuntimeError("Table 1의 사다리 사용 작업 법령 보완값이 예상과 다릅니다.")
    return summary


def build_live_verification_summary(
    imported_rows: list[ImportRow],
    fetched_items: list[dict[str, Any]],
) -> dict[str, Any]:
    imported_by_code = {
        normalize_cell_text(item.get("code")): item
        for item in fetched_items
        if normalize_cell_text(item.get("code")).startswith("hazard-cm-")
        and item.get("is_active") is True
    }

    table1_items = [
        item for code, item in imported_by_code.items() if code.startswith("hazard-cm-table1-")
    ]
    sheet1_items = [
        item for code, item in imported_by_code.items() if code.startswith("hazard-cm-sheet1-")
    ]

    if len(imported_by_code) != EXPECTED_TOTAL_ROWS:
        raise RuntimeError(
            "활성 hazard-cm 코드 건수가 예상과 다릅니다. "
            f"expected={EXPECTED_TOTAL_ROWS}, actual={len(imported_by_code)}"
        )
    if len(table1_items) != REQUIRED_ROW_COUNT_PER_SHEET:
        raise RuntimeError("활성 Table 1 코드 건수가 161이 아닙니다.")
    if len(sheet1_items) != REQUIRED_ROW_COUNT_PER_SHEET:
        raise RuntimeError("활성 Sheet1 코드 건수가 161이 아닙니다.")

    category_nonempty_count = 0
    for item in imported_by_code.values():
        body = item.get("body")
        if isinstance(body, dict) and normalize_cell_text(body.get("category")):
            category_nonempty_count += 1
    if category_nonempty_count != EXPECTED_SHEET1_NONEMPTY_CATEGORY_COUNT:
        raise RuntimeError(
            "활성 데이터의 non-empty category 건수가 예상과 다릅니다. "
            f"expected={EXPECTED_SHEET1_NONEMPTY_CATEGORY_COUNT}, actual={category_nonempty_count}"
        )

    ladder_item = imported_by_code.get("hazard-cm-table1-012")
    if not ladder_item:
        raise RuntimeError("hazard-cm-table1-012를 검증용 조회 결과에서 찾지 못했습니다.")
    ladder_body = ladder_item.get("body")
    ladder_legal_reference = ""
    if isinstance(ladder_body, dict):
        ladder_legal_reference = normalize_cell_text(ladder_body.get("legalReference"))
    if ladder_legal_reference != "산업안전보건기준에 관한 규칙 제42조(추락의 방지)":
        raise RuntimeError("hazard-cm-table1-012 법령 값이 예상과 다릅니다.")

    sample_codes = [
        "hazard-cm-table1-001",
        "hazard-cm-table1-002",
        "hazard-cm-table1-003",
        "hazard-cm-table1-004",
        "hazard-cm-table1-005",
        "hazard-cm-sheet1-001",
        "hazard-cm-sheet1-002",
        "hazard-cm-sheet1-003",
        "hazard-cm-sheet1-004",
        "hazard-cm-sheet1-005",
    ]
    imported_row_map = {row.code: row for row in imported_rows}
    samples: list[dict[str, Any]] = []
    for code in sample_codes:
        expected_row = imported_row_map[code]
        actual_item = imported_by_code.get(code)
        if not actual_item:
            raise RuntimeError(f"샘플 코드 {code}를 서버에서 찾지 못했습니다.")
        body = actual_item.get("body")
        if not isinstance(body, dict):
            raise RuntimeError(f"샘플 코드 {code}의 body가 객체 형식이 아닙니다.")

        actual_snapshot = {
            "title": normalize_cell_text(actual_item.get("title")),
            "category": normalize_cell_text(body.get("category")),
            "expectedRisk": normalize_cell_text(body.get("expectedRisk")),
            "countermeasure": normalize_cell_text(body.get("countermeasure")),
            "legalReference": normalize_cell_text(body.get("legalReference")),
            "note": normalize_cell_text(body.get("note")),
        }
        expected_snapshot = {
            "title": expected_row.title,
            "category": expected_row.category,
            "expectedRisk": expected_row.expected_risk,
            "countermeasure": expected_row.countermeasure,
            "legalReference": expected_row.legal_reference,
            "note": expected_row.note,
        }
        if actual_snapshot != expected_snapshot:
            raise RuntimeError(
                f"샘플 코드 {code}의 서버 데이터가 원본과 다릅니다.\n"
                f"expected={expected_snapshot}\nactual={actual_snapshot}"
            )
        samples.append({"code": code, "title": expected_row.title})

    return {
        "active_imported_count": len(imported_by_code),
        "active_table1_count": len(table1_items),
        "active_sheet1_count": len(sheet1_items),
        "nonempty_category_count": category_nonempty_count,
        "ladder_legal_reference": ladder_legal_reference,
        "verified_samples": samples,
    }


def print_json(data: dict[str, Any]) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(
        description="요인과 대책.xlsx를 예상위험&관리대책(hazard_countermeasure_catalog)으로 적재합니다."
    )
    parser.add_argument("workbook_path", type=Path, help="업로드할 xlsx 파일 경로")
    parser.add_argument(
        "--api-base",
        dest="api_base",
        default=None,
        help="Safety API upstream base URL. 기본값은 현재 upstream 규칙을 따릅니다.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="서버에 쓰지 않고 변환/검증 요약만 출력합니다.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise RuntimeError(f"워크북 파일을 찾지 못했습니다: {workbook_path}")

    api_base_url = resolve_api_base_url(args.api_base)
    import_rows = build_import_rows(workbook_path)
    validation_summary = validate_import_rows(import_rows)

    base_summary = {
        "mode": "dry-run" if args.dry_run else "live",
        "workbook_path": str(workbook_path),
        "api_base_url": api_base_url,
        "validation": validation_summary,
    }

    if args.dry_run:
        print_json(base_summary)
        return

    access_token = resolve_access_token(api_base_url)
    all_existing_items = fetch_all_content_items(
        api_base_url,
        access_token,
        active_only=False,
        include_body=False,
    )
    active_hazard_items = [
        item
        for item in all_existing_items
        if item.get("content_type") == "hazard_countermeasure_catalog"
        and item.get("is_active") is True
    ]
    if active_hazard_items:
        raise RuntimeError(
            "활성 hazard_countermeasure_catalog 데이터가 이미 존재합니다. "
            f"expected=0, actual={len(active_hazard_items)}"
        )

    created_codes: list[str] = []
    for index, row in enumerate(import_rows, start=1):
        post_content_item(api_base_url, access_token, row.to_payload())
        created_codes.append(row.code)
        if index % 25 == 0 or index == len(import_rows):
            print(f"[import] created {index}/{len(import_rows)}")

    post_import_items = fetch_all_content_items(
        api_base_url,
        access_token,
        active_only=False,
        include_body=True,
    )
    verification_summary = build_live_verification_summary(import_rows, post_import_items)

    print_json(
        {
            **base_summary,
            "created_count": len(created_codes),
            "verification": verification_summary,
        }
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:  # noqa: BLE001
        message = str(error) or error.__class__.__name__
        print(f"import_hazard_countermeasure_catalog failed: {message}", file=sys.stderr)
        sys.exit(1)
