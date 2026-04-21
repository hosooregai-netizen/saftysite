#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import sys
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_SAFETY_API_UPSTREAM_BASE_URL = "http://52.64.85.49:8011/api/v1"
API_BASE_ENV_KEYS = (
    "SAFETY_API_UPSTREAM_BASE_URL",
    "SAFETY_API_BASE_URL",
    "NEXT_PUBLIC_SAFETY_API_UPSTREAM_BASE_URL",
    "NEXT_PUBLIC_SAFETY_API_BASE_URL",
)
CONTENT_LIST_LIMIT = 500
GENERAL_CAUSATIVE = "일반"
LEFT_BLOCK = (2, 3, 4, "left")
RIGHT_BLOCK = (9, 10, 11, "right")
NUMBERED_FILE_PATTERN = re.compile(r"^\s*(\d+)")
SHEET_SLUGS = {
    "추락": "fall",
    "끼임": "caught-in",
    "부딪힘": "struck",
    "깔림": "crushed",
    "맞음": "hit",
    "기타": "other",
}
SPECIAL_CAUSATIVE_OVERRIDES = {
    ("부딪힘", 9): "지게차",
}


@dataclass(frozen=True)
class ImportEntry:
    sheet_name: str
    sheet_slug: str
    number: int
    source_row_number: int
    side: str
    source_title: str
    accident_type: str
    causative_agent: str
    body: str
    image_path: Path
    code: str
    sort_order: int

    @property
    def title(self) -> str:
        return f"{self.accident_type} / {self.causative_agent}"

    @property
    def tags(self) -> list[str]:
        return ["excel-import", "doc7-reference-material", self.sheet_slug]

    def to_payload(self, image_url: str) -> dict[str, Any]:
        return {
            "content_type": "doc7_reference_material",
            "title": self.title,
            "code": self.code,
            "body": {
                "accidentType": self.accident_type,
                "causativeAgentKey": self.causative_agent,
                "body": self.body,
                "imageUrl": image_url,
            },
            "tags": self.tags,
            "sort_order": self.sort_order,
            "effective_from": None,
            "effective_to": None,
            "is_active": True,
        }


@dataclass
class ExistingDoc7Item:
    id: str
    title: str
    code: str
    accident_type: str
    causative_agent: str
    body: str
    image_url: str
    sort_order: int
    is_active: bool
    tags: tuple[str, ...]


def normalize_base_url(value: str) -> str:
    return value.rstrip("/")


def is_absolute_http_url(value: str) -> bool:
    return bool(re.match(r"^https?://", value, flags=re.I))


def resolve_api_base_url(explicit: str | None) -> str:
    if explicit and is_absolute_http_url(explicit.strip()):
        return normalize_base_url(explicit.strip())

    for env_key in API_BASE_ENV_KEYS:
        configured = os.getenv(env_key, "").strip()
        if configured and is_absolute_http_url(configured):
            return normalize_base_url(configured)

    return DEFAULT_SAFETY_API_UPSTREAM_BASE_URL


def normalize_line_endings(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def normalize_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_line_endings(value).strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_line_endings(str(value)).strip()


def collapse_whitespace(value: str) -> str:
    return " ".join(normalize_line_endings(value).split())


def normalize_match_value(value: str) -> str:
    return re.sub(r"[\s\-_/,()]+", "", value).lower()


def extract_number_from_filename(path: Path) -> int | None:
    match = NUMBERED_FILE_PATTERN.match(path.stem)
    return int(match.group(1)) if match else None


def as_record(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def build_content_image_url(body: Any) -> str:
    record = as_record(body)
    return (
        normalize_cell_text(record.get("imageUrl"))
        or normalize_cell_text(record.get("image_url"))
        or normalize_cell_text(record.get("fileUrl"))
        or normalize_cell_text(record.get("file_url"))
        or normalize_cell_text(record.get("assetUrl"))
        or normalize_cell_text(record.get("asset_url"))
        or normalize_cell_text(record.get("thumbnailUrl"))
        or normalize_cell_text(record.get("thumbnail_url"))
    )


def read_doc7_body_text(body: Any) -> str:
    record = as_record(body)
    return (
        normalize_cell_text(record.get("body"))
        or normalize_cell_text(record.get("summary"))
        or normalize_cell_text(record.get("description"))
        or normalize_cell_text(record.get("content"))
        or normalize_cell_text(record.get("text"))
    )


def parse_title_fields(sheet_name: str, number: int, title: str) -> tuple[str, str]:
    normalized_title = collapse_whitespace(title)
    match = re.match(r"^(.*?)\s*\((.*?)\)\s*$", normalized_title)
    if match:
        accident_type = collapse_whitespace(match.group(1))
        causative_agent = collapse_whitespace(match.group(2))
    else:
        accident_type = normalized_title
        causative_agent = GENERAL_CAUSATIVE

    override = SPECIAL_CAUSATIVE_OVERRIDES.get((sheet_name, number))
    if override:
        causative_agent = override

    return accident_type, causative_agent


def request_json(
    method: str,
    url: str,
    *,
    headers: dict[str, str] | None = None,
    body: bytes | None = None,
    timeout: int = 60,
) -> Any:
    request = Request(url, data=body, method=method.upper())
    for key, value in (headers or {}).items():
        request.add_header(key, value)

    try:
        with urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
            if not raw.strip():
                return None
            return json.loads(raw)
    except HTTPError as error:
        raw = error.read().decode("utf-8", errors="replace")
        detail = raw
        if raw:
            try:
                payload = json.loads(raw)
                if isinstance(payload, dict) and isinstance(payload.get("detail"), str):
                    detail = payload["detail"]
                else:
                    detail = json.dumps(payload, ensure_ascii=False)
            except json.JSONDecodeError:
                detail = raw
        raise RuntimeError(f"{method.upper()} {url} failed ({error.code}): {detail}") from error
    except URLError as error:
        raise RuntimeError(f"{method.upper()} {url} failed: {error.reason}") from error


def encode_multipart_file(field_name: str, path: Path) -> tuple[bytes, str]:
    boundary = f"----Doc7Reference{uuid.uuid4().hex}"
    file_bytes = path.read_bytes()
    content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    parts = [
        f"--{boundary}\r\n".encode("utf-8"),
        (
            f'Content-Disposition: form-data; name="{field_name}"; filename="{path.name}"\r\n'
        ).encode("utf-8"),
        f"Content-Type: {content_type}\r\n\r\n".encode("utf-8"),
        file_bytes,
        b"\r\n",
        f"--{boundary}--\r\n".encode("utf-8"),
    ]
    body = b"".join(parts)
    return body, f"multipart/form-data; boundary={boundary}"


def resolve_access_token(api_base_url: str) -> str:
    access_token = os.getenv("SAFETY_ACCESS_TOKEN", "").strip()
    if access_token:
        return access_token

    email = (
        os.getenv("SAFETY_ADMIN_EMAIL", "").strip()
        or os.getenv("LIVE_SAFETY_EMAIL", "").strip()
    )
    password = os.getenv("SAFETY_ADMIN_PASSWORD") or os.getenv("LIVE_SAFETY_PASSWORD")
    if not email or password is None:
        raise RuntimeError(
            "SAFETY_ACCESS_TOKEN 또는 SAFETY_ADMIN_EMAIL / SAFETY_ADMIN_PASSWORD 환경변수가 필요합니다."
        )

    body = urlencode({"username": email, "password": password}).encode("utf-8")
    response = request_json(
        "POST",
        f"{api_base_url}/auth/token",
        headers={"Content-Type": "application/x-www-form-urlencoded; charset=utf-8"},
        body=body,
    )
    token = response.get("access_token") if isinstance(response, dict) else None
    if not isinstance(token, str) or not token.strip():
        raise RuntimeError("인증 응답에서 access_token을 찾지 못했습니다.")
    return token.strip()


def fetch_all_content_items(
    api_base_url: str,
    access_token: str,
    *,
    active_only: bool,
    include_body: bool,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    offset = 0
    headers = {"Authorization": f"Bearer {access_token}"}

    while True:
        query = urlencode(
            {
                "active_only": str(active_only).lower(),
                "include_body": str(include_body).lower(),
                "limit": CONTENT_LIST_LIMIT,
                "offset": offset,
            }
        )
        page = request_json(
            "GET",
            f"{api_base_url}/content-items?{query}",
            headers=headers,
        )
        if not isinstance(page, list):
            raise RuntimeError("content-items 응답 형식이 예상과 다릅니다.")

        items.extend(page)
        if len(page) < CONTENT_LIST_LIMIT:
            return items
        offset += len(page)


def upload_content_asset(
    api_base_url: str,
    access_token: str,
    image_path: Path,
) -> str:
    body, content_type = encode_multipart_file("file", image_path)
    response = request_json(
        "POST",
        f"{api_base_url}/content-items/assets/upload",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": content_type,
        },
        body=body,
        timeout=90,
    )
    if not isinstance(response, dict):
        raise RuntimeError("이미지 업로드 응답 형식이 예상과 다릅니다.")

    uploaded_url = normalize_cell_text(response.get("url")) or normalize_cell_text(
        response.get("path")
    )
    if not uploaded_url:
        raise RuntimeError(f"이미지 업로드 URL이 비어 있습니다: {image_path}")
    return uploaded_url


def post_content_item(
    api_base_url: str,
    access_token: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    response = request_json(
        "POST",
        f"{api_base_url}/content-items",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        body=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
    )
    if not isinstance(response, dict):
        raise RuntimeError("content-items 생성 응답 형식이 예상과 다릅니다.")
    return response


def patch_content_item(
    api_base_url: str,
    access_token: str,
    item_id: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    response = request_json(
        "PATCH",
        f"{api_base_url}/content-items/{item_id}",
        headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        body=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
    )
    if not isinstance(response, dict):
        raise RuntimeError("content-items 수정 응답 형식이 예상과 다릅니다.")
    return response


def build_numbered_image_map(folder: Path) -> dict[int, Path]:
    numbered_images: dict[int, Path] = {}
    for path in folder.iterdir():
        if not path.is_file():
            continue
        number = extract_number_from_filename(path)
        if number is None:
            continue
        if number in numbered_images:
            raise RuntimeError(f"{folder} 에 번호가 중복된 이미지가 있습니다: {number}")
        numbered_images[number] = path
    return numbered_images


def detect_image_root(start_dir: Path, sheet_names: list[str]) -> Path:
    candidates = [start_dir]
    candidates.extend([path for path in start_dir.iterdir() if path.is_dir()])

    for candidate in candidates:
        if not candidate.is_dir():
            continue
        child_dirs = {path.name: path for path in candidate.iterdir() if path.is_dir()}
        if all(sheet_name in child_dirs for sheet_name in sheet_names):
            if all(
                any(file.suffix.lower() == ".png" for file in child_dirs[sheet_name].iterdir())
                for sheet_name in sheet_names
            ):
                return candidate

    raise RuntimeError(
        "참고자료 이미지 루트를 자동으로 찾지 못했습니다. --image-root 로 직접 지정해주세요."
    )


def build_import_entries(workbook_path: Path, image_root: Path) -> list[ImportEntry]:
    workbook = load_workbook(workbook_path, data_only=True)
    entries: list[ImportEntry] = []
    sort_order = 0

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        sheet_slug = SHEET_SLUGS.get(sheet_name)
        if not sheet_slug:
            continue

        image_folder = image_root / sheet_name
        if not image_folder.is_dir():
            raise RuntimeError(f"시트 {sheet_name} 에 대응하는 이미지 폴더가 없습니다: {image_folder}")
        numbered_images = build_numbered_image_map(image_folder)

        for row_number in range(1, worksheet.max_row + 1):
            for number_col, title_col, body_col, side in (LEFT_BLOCK, RIGHT_BLOCK):
                raw_number = worksheet.cell(row_number, number_col).value
                title = normalize_cell_text(worksheet.cell(row_number, title_col).value)
                body = normalize_cell_text(worksheet.cell(row_number, body_col).value)
                if not title:
                    continue
                if not isinstance(raw_number, (int, float)) or int(raw_number) != raw_number:
                    continue

                number = int(raw_number)
                image_path = numbered_images.get(number)
                if image_path is None:
                    raise RuntimeError(
                        f"{sheet_name} #{number} 에 대응하는 이미지가 없습니다. 폴더={image_folder}"
                    )

                accident_type, causative_agent = parse_title_fields(sheet_name, number, title)
                entries.append(
                    ImportEntry(
                        sheet_name=sheet_name,
                        sheet_slug=sheet_slug,
                        number=number,
                        source_row_number=row_number,
                        side=side,
                        source_title=collapse_whitespace(title),
                        accident_type=accident_type,
                        causative_agent=causative_agent,
                        body=body,
                        image_path=image_path,
                        code=f"doc7-ref-{sheet_slug}-{number:03d}",
                        sort_order=sort_order,
                    )
                )
                sort_order += 1

        if len(numbered_images) != len([entry for entry in entries if entry.sheet_name == sheet_name]):
            raise RuntimeError(
                f"{sheet_name} 시트 건수와 이미지 건수가 일치하지 않습니다. "
                f"sheet={len([entry for entry in entries if entry.sheet_name == sheet_name])}, "
                f"images={len(numbered_images)}"
            )

    if not entries:
        raise RuntimeError("가져올 DOC7 참고자료를 한 건도 찾지 못했습니다.")

    return entries


def map_existing_doc7_items(items: list[dict[str, Any]]) -> list[ExistingDoc7Item]:
    mapped: list[ExistingDoc7Item] = []

    for item in items:
        if normalize_cell_text(item.get("content_type")) != "doc7_reference_material":
            continue

        body = as_record(item.get("body"))
        mapped.append(
            ExistingDoc7Item(
                id=normalize_cell_text(item.get("id")),
                title=collapse_whitespace(normalize_cell_text(item.get("title"))),
                code=normalize_cell_text(item.get("code")),
                accident_type=collapse_whitespace(
                    normalize_cell_text(body.get("accidentType") or body.get("accident_type"))
                ),
                causative_agent=collapse_whitespace(
                    normalize_cell_text(body.get("causativeAgentKey") or body.get("causative_agent_key"))
                ),
                body=read_doc7_body_text(body),
                image_url=build_content_image_url(body),
                sort_order=int(item.get("sort_order") or 0),
                is_active=bool(item.get("is_active")),
                tags=tuple(
                    sorted(
                        normalize_cell_text(tag)
                        for tag in item.get("tags", [])
                        if normalize_cell_text(tag)
                    )
                ),
            )
        )

    return mapped


def build_entry_match_key(accident_type: str, causative_agent: str) -> str:
    return f"{normalize_match_value(accident_type)}|{normalize_match_value(causative_agent)}"


def find_existing_match(
    entry: ImportEntry,
    existing_items: list[ExistingDoc7Item],
    matched_ids: set[str],
) -> ExistingDoc7Item | None:
    by_code = next(
        (
            item
            for item in existing_items
            if item.id not in matched_ids and item.code == entry.code
        ),
        None,
    )
    if by_code:
        return by_code

    entry_title = build_entry_match_key(entry.accident_type, entry.causative_agent)
    by_pair = next(
        (
            item
            for item in existing_items
            if item.id not in matched_ids
            and build_entry_match_key(item.accident_type, item.causative_agent) == entry_title
        ),
        None,
    )
    if by_pair:
        return by_pair

    normalized_title = build_entry_match_key(entry.title, "")
    return next(
        (
            item
            for item in existing_items
            if item.id not in matched_ids
            and build_entry_match_key(item.title, "") == normalized_title
        ),
        None,
    )


def needs_update(existing: ExistingDoc7Item, entry: ImportEntry) -> bool:
    return any(
        [
            existing.code != entry.code,
            collapse_whitespace(existing.title) != entry.title,
            build_entry_match_key(existing.accident_type, existing.causative_agent)
            != build_entry_match_key(entry.accident_type, entry.causative_agent),
            normalize_line_endings(existing.body) != normalize_line_endings(entry.body),
            existing.sort_order != entry.sort_order,
            not existing.is_active,
            existing.tags != tuple(sorted(entry.tags)),
            not existing.image_url,
        ]
    )


def summarize_dry_run(
    entries: list[ImportEntry],
    existing_items: list[ExistingDoc7Item],
) -> dict[str, Any]:
    matched_ids: set[str] = set()
    create_count = 0
    update_count = 0
    skip_count = 0
    samples: list[dict[str, Any]] = []

    for entry in entries:
        existing = find_existing_match(entry, existing_items, matched_ids)
        if existing is None:
            create_count += 1
            action = "create"
        else:
            matched_ids.add(existing.id)
            if needs_update(existing, entry):
                update_count += 1
                action = "update"
            else:
                skip_count += 1
                action = "skip"

        if len(samples) < 12:
            samples.append(
                {
                    "action": action,
                    "code": entry.code,
                    "sheet": entry.sheet_name,
                    "number": entry.number,
                    "title": entry.title,
                    "image": entry.image_path.name,
                }
            )

    unmatched_existing = [
        item
        for item in existing_items
        if item.id not in matched_ids
    ]

    sheet_counts: dict[str, int] = {}
    for entry in entries:
        sheet_counts[entry.sheet_name] = sheet_counts.get(entry.sheet_name, 0) + 1

    return {
        "total_import_entries": len(entries),
        "existing_doc7_reference_materials": len(existing_items),
        "create_count": create_count,
        "update_count": update_count,
        "skip_count": skip_count,
        "unmatched_existing_count": len(unmatched_existing),
        "sheet_counts": sheet_counts,
        "samples": samples,
        "unmatched_existing_titles_preview": [item.title for item in unmatched_existing[:10]],
    }


def verify_post_import(
    entries: list[ImportEntry],
    refreshed_items: list[ExistingDoc7Item],
) -> dict[str, Any]:
    refreshed_by_code = {item.code: item for item in refreshed_items if item.code}
    missing_codes = [entry.code for entry in entries if entry.code not in refreshed_by_code]
    if missing_codes:
        raise RuntimeError(f"삽입 후 코드가 누락된 참고자료가 있습니다: {missing_codes[:10]}")

    wrong_pairs: list[str] = []
    missing_images: list[str] = []
    for entry in entries:
        actual = refreshed_by_code[entry.code]
        if build_entry_match_key(actual.accident_type, actual.causative_agent) != build_entry_match_key(
            entry.accident_type,
            entry.causative_agent,
        ):
            wrong_pairs.append(entry.code)
        if not actual.image_url:
            missing_images.append(entry.code)

    if wrong_pairs:
        raise RuntimeError(f"삽입 후 사고유형/기인물 매핑이 어긋난 항목이 있습니다: {wrong_pairs[:10]}")
    if missing_images:
        raise RuntimeError(f"삽입 후 이미지 URL이 비어 있는 항목이 있습니다: {missing_images[:10]}")

    return {
        "verified_count": len(entries),
        "sample_codes": [entry.code for entry in entries[:10]],
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="재해사례 엑셀과 참고자료 이미지를 DOC7 참고자료 콘텐츠로 업서트합니다."
    )
    parser.add_argument(
        "workbook_path",
        type=Path,
        help="재해사례 정리.xlsx 경로 또는 드라이런용 임시 복사본 경로",
    )
    parser.add_argument(
        "--image-root",
        type=Path,
        default=None,
        help="참고자료 이미지 루트. 생략 시 현재 작업 디렉터리 기준 자동 탐지합니다.",
    )
    parser.add_argument(
        "--api-base",
        dest="api_base",
        default=None,
        help="Safety API upstream base URL. 생략 시 환경변수 또는 기본 upstream을 사용합니다.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="서버에 반영하지 않고 매핑/업서트 요약만 출력합니다.",
    )
    args = parser.parse_args()

    workbook_path = args.workbook_path.expanduser().resolve()
    if not workbook_path.exists():
        raise RuntimeError(f"엑셀 파일을 찾지 못했습니다: {workbook_path}")

    preview_workbook = load_workbook(workbook_path, data_only=True)
    sheet_names = [worksheet.title for worksheet in preview_workbook.worksheets if worksheet.title in SHEET_SLUGS]
    if not sheet_names:
        raise RuntimeError("지원하는 시트(추락/끼임/부딪힘/깔림/맞음/기타)를 찾지 못했습니다.")

    image_root = (
        args.image_root.expanduser().resolve()
        if args.image_root is not None
        else detect_image_root(Path.cwd(), sheet_names)
    )
    if not image_root.is_dir():
        raise RuntimeError(f"이미지 루트를 찾지 못했습니다: {image_root}")

    entries = build_import_entries(workbook_path, image_root)
    api_base_url = resolve_api_base_url(args.api_base)
    access_token = resolve_access_token(api_base_url)
    existing_raw_items = fetch_all_content_items(
        api_base_url,
        access_token,
        active_only=False,
        include_body=True,
    )
    existing_doc7_items = map_existing_doc7_items(existing_raw_items)

    if args.dry_run:
        payload = {
            "api_base_url": api_base_url,
            "image_root": str(image_root),
            "workbook_path": str(workbook_path),
            "summary": summarize_dry_run(entries, existing_doc7_items),
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    matched_ids: set[str] = set()
    created_codes: list[str] = []
    updated_codes: list[str] = []
    skipped_codes: list[str] = []

    total = len(entries)
    for index, entry in enumerate(entries, start=1):
        existing = find_existing_match(entry, existing_doc7_items, matched_ids)
        if existing is not None:
            matched_ids.add(existing.id)
            if not needs_update(existing, entry):
                skipped_codes.append(entry.code)
                print(f"[{index}/{total}] skip {entry.code} {entry.title}")
                continue

        uploaded_url = upload_content_asset(api_base_url, access_token, entry.image_path)
        payload = entry.to_payload(uploaded_url)

        if existing is None:
            post_content_item(api_base_url, access_token, payload)
            created_codes.append(entry.code)
            print(f"[{index}/{total}] create {entry.code} {entry.title}")
            continue

        patch_content_item(
            api_base_url,
            access_token,
            existing.id,
            {
                "title": payload["title"],
                "code": payload["code"],
                "body": payload["body"],
                "tags": payload["tags"],
                "sort_order": payload["sort_order"],
                "effective_from": payload["effective_from"],
                "effective_to": payload["effective_to"],
                "is_active": payload["is_active"],
            },
        )
        updated_codes.append(entry.code)
        print(f"[{index}/{total}] update {entry.code} {entry.title}")

    refreshed_raw_items = fetch_all_content_items(
        api_base_url,
        access_token,
        active_only=False,
        include_body=True,
    )
    refreshed_doc7_items = map_existing_doc7_items(refreshed_raw_items)
    verification = verify_post_import(entries, refreshed_doc7_items)

    print(
        json.dumps(
            {
                "api_base_url": api_base_url,
                "created_count": len(created_codes),
                "updated_count": len(updated_codes),
                "skipped_count": len(skipped_codes),
                "created_codes": created_codes,
                "updated_codes": updated_codes,
                "skipped_codes": skipped_codes,
                "verification": verification,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:  # noqa: BLE001
        print(f"DOC7 reference import failed: {error}", file=sys.stderr)
        sys.exit(1)
